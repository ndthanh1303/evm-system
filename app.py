import MySQLdb.cursors
import re
import os
import io
import pandas as pd
from io import BytesIO
from flask import (
    Flask,
    render_template,
    request,
    redirect,
    send_file,
    url_for,
    session,
    flash
)

from flask_mysqldb import MySQL
from flask_mail import Mail, Message
from datetime import timedelta
from datetime import datetime

from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from openpyxl.styles import (
    Font,
    Alignment,
    PatternFill,
    Border,
    Side
)
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

app = Flask(__name__)

# ======================
# MAIL CONFIG
# ======================

app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True

# EMAIL GỬI
app.config['MAIL_USERNAME'] = 'ndthanh.26102016@gmail.com'

# APP PASSWORD
app.config['MAIL_PASSWORD'] = 'iiandlnocywkoxjs'

app.config['MAIL_DEFAULT_SENDER'] = 'ndthanh.26102016@gmail.com'

mail = Mail(app)

# 🔐 Secret key để dùng session
app.secret_key = '123456'

# ⚙️ Cấu hình MySQL
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = '123456'
app.config['MYSQL_DB'] = 'evm_system'

mysql = MySQL(app)

# ===== GLOBAL USER =====
@app.context_processor
def inject_user():
    return dict(
        fullname=session.get('fullname')
    )

# app.py
def parse_number(val):
    return float(val.replace(",", "")) if val else 0

#Home
@app.route('/')
def home():
    if 'loggedin' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

# =========================
# Đăng nhập
# ========================
@app.route('/login', methods=['GET', 'POST'])
def login():
    msg = ''

    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        remember = request.form.get('remember')  # None nếu không tick

        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute(
            'SELECT * FROM users WHERE username = %s AND password = %s',
            (username, password)
        )
        account = cursor.fetchone()

        if account:
            session['loggedin'] = True
            session['id'] = account['user_id']
            session['username'] = account['username']
            session['fullname'] = account['fullname']
            session['role'] = account['role']

            # 
            if remember:
                session.permanent = True
                app.permanent_session_lifetime = timedelta(days=7)
            else:
                session.permanent = False

            return redirect(url_for('dashboard'))
        else:
            msg = 'Sai tài khoản hoặc mật khẩu!'

    return render_template('login.html', msg=msg)

# =========================
# Đăng ký
# =========================
@app.route('/register', methods=['GET', 'POST'])
def register():
    msg = ''
    fullname = ''
    phone = ''
    email = ''
    username = ''

    if request.method == 'POST':
        fullname = request.form['fullname']
        phone = request.form['phone']
        email = request.form['email']
        username = request.form['username']
        password = request.form['password']
        repassword = request.form['repassword']

        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

        # 1. Kiểm tra đủ dữ liệu
        if not fullname or not phone or not email or not username or not password or not repassword:
            msg = 'Vui lòng nhập đầy đủ thông tin!'

        # 2. Kiểm tra password trùng nhau
        elif password != repassword:
            msg = 'Password không trùng nhau!'

        elif not re.match(r"^(0|\+84)[0-9]{9,10}$", phone):
            msg = "Số điện thoại không hợp lệ!"

        elif not re.match(r"^[\w\.-]+@[\w\.-]+\.\w+$", email):
            msg = "Email không hợp lệ!"

        else:
            # 3. Kiểm tra username tồn tại
            cursor.execute('SELECT * FROM users WHERE username = %s', (username,))
            account = cursor.fetchone()

            # 4. Kiểm tra email tồn tại
            cursor.execute('SELECT * FROM users WHERE email = %s', (email,))
            email_exist = cursor.fetchone()

            if account:
                msg = 'Username đã tồn tại!'
            elif email_exist:
                msg = 'Email đã được sử dụng!'
            else:
                # Thêm vào database
                cursor.execute(
                    '''INSERT INTO users (fullname, phone, email, username, password)
                       VALUES (%s, %s, %s, %s, %s)''',
                    (fullname, phone, email, username, password)
                )
                mysql.connection.commit()
                msg = 'Đăng ký thành công!'
                return render_template('register.html',msg=msg)

    return render_template(
            'register.html',
            msg=msg,
            fullname=fullname,
            phone=phone,
            email=email,
            username=username
    )

# =========================
# Dashboard (sau login)
# =========================
@app.route('/dashboard')
def dashboard():

    if 'loggedin' not in session:
        return redirect(url_for('login'))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    user_id = session['id']
    role = session['role']


    #========================= LẤY PROJECT =========================
    if role == 'admin':
        cursor.execute("SELECT * FROM projects")
    else:
        cursor.execute(
            "SELECT * FROM projects WHERE user_id = %s",
            (user_id,)
        )

    projects = cursor.fetchall()

    #========================= THỐNG KÊ STATUS =========================
    if role == 'admin':
        cursor.execute("""
            SELECT status, COUNT(*) as total
            FROM projects
            GROUP BY status
        """)
    else:
        cursor.execute("""
            SELECT status, COUNT(*) as total
            FROM projects
            WHERE user_id = %s
            GROUP BY status
        """, (user_id,))

    rows = cursor.fetchall()

    status_map = {
        'ke_hoach': 0,
        'dang_thuc_hien': 0,
        'tam_dung': 0,
        'hoan_thanh': 0,
        'dong': 0
    }

    for r in rows:
        status_map[r['status']] = r['total']

    planning_count = status_map.get('ke_hoach', 0)
    working_count = status_map.get('dang_thuc_hien', 0)
    pause_count = status_map.get('tam_dung', 0)
    done_count = status_map.get('hoan_thanh', 0)
    close_count = status_map.get('dong', 0)

    # ========================= TÍNH SPI / CPI =========================
    late_projects = []
    over_budget_projects = []

    for p in projects:

        cursor.execute("""
            SELECT 
                SUM(PV_kluong * PV_dongia) AS PV,
                SUM(EV_kluong * EV_dongia) AS EV,
                SUM(AC_kluong * AC_dongia) AS AC
            FROM work
            WHERE proj_id = %s
        """, (p['proj_id'],))

        data = cursor.fetchone()

        PV = data['PV'] or 0
        EV = data['EV'] or 0
        AC = data['AC'] or 0

        SPI = EV / PV if PV > 0 else 0
        CPI = EV / AC if AC > 0 else 0

        # ========================= TRỄ TIẾN ĐỘ =========================
        if SPI < 1:
            late_projects.append({
                'proj_name': p['proj_name']
            })

        # ========================= VƯỢT CHI PHÍ =========================
        if CPI < 1:
            over_budget_projects.append({
                'proj_name': p['proj_name']
            })

    #========================= THỐNG KÊ BIỂU ĐỒ =========================
    total_projects = len(projects)

    late_count = len(late_projects)
    on_time = total_projects - late_count

    over_budget_count = len(over_budget_projects)
    saving_budget_count = total_projects - over_budget_count

    #========================= RENDER =========================
    return render_template(
        'dashboard.html',

        username=session['username'],
        active_page='dashboard',

        # STATUS
        planning_count=planning_count,
        working_count=working_count,
        pause_count=pause_count,
        done_count=done_count,
        close_count=close_count,

        # PROGRESS
        late_count=late_count,
        on_time=on_time,

        # COST
        over_budget_count=over_budget_count,
        saving_budget_count=saving_budget_count,

        # LIST
        late_projects=late_projects,
        over_budget_projects=over_budget_projects
    )

# ======================
# PROJECT
# ======================
@app.route('/project')
def project():
    # ===== CHECK LOGIN =====
    if 'loggedin' not in session:
        return redirect(url_for('login'))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    filter_type = request.args.get('filter_type')
    date_type = request.args.get('date_type')

    if not date_type:
        date_type = 'created_at'

    # ===== BASE QUERY =====
    query = "SELECT * FROM projects WHERE 1=1"
    params = []

    # ===== PHÂN QUYỀN =====
    if session['role'] != 'admin':
        query += " AND user_id = %s"
        params.append(session['id'])

    # ===== FILTER STATUS =====
    status = request.args.get('status')
    if status:
        query += " AND status = %s"
        params.append(status)

    # ===== search =====
    keyword = request.args.get('keyword')
    if keyword:
        query += " AND proj_name LIKE %s"
        params.append(f"%{keyword}%")
        
    # ===== THEO NGÀY =====
    if filter_type == 'day':
        date = request.args.get('date')
        if date:
            query += f" AND DATE({date_type}) = %s"
            params.append(date)

    # ===== THEO THÁNG =====
    elif filter_type == 'month':
        month = request.args.get('month')
        year = request.args.get('year')

        if month and year:
            query += f"""
                AND MONTH({date_type}) = %s
                AND YEAR({date_type}) = %s
            """
            params.extend([month, year])

    # ===== THEO QUÝ =====
    elif filter_type == 'quarter':
        quarter = request.args.get('quarter')
        year = request.args.get('year')

        if quarter and year:
            quarter = int(quarter)

            if quarter == 1:
                start_month, end_month = 1, 3
            elif quarter == 2:
                start_month, end_month = 4, 6
            elif quarter == 3:
                start_month, end_month = 7, 9
            else:
                start_month, end_month = 10, 12

            query += f"""
                AND MONTH({date_type}) BETWEEN %s AND %s
                AND YEAR({date_type}) = %s
            """
            params.extend([start_month, end_month, year])

    # ===== THEO NĂM =====
    elif filter_type == 'year':
        year = request.args.get('year')

        if year:
            query += f" AND YEAR({date_type}) = %s"
            params.append(year)

    # ===== EXECUTE =====
    cursor.execute(query, tuple(params))
    projects = cursor.fetchall()

        
    # ===== THỐNG KÊ =====
    stats = {
        'ke_hoach': 0,
        'dang_thuc_hien': 0,
        'tam_dung': 0,
        'hoan_thanh': 0,
        'dong': 0
    }

    for p in projects:
        stats[p['status']] += 1

    status_map = {
        'ke_hoach': 'Kế hoạch',
        'dang_thuc_hien': 'Đang thực hiện',
        'tam_dung': 'Tạm dừng',
        'hoan_thanh': 'Hoàn thành',
        'dong': 'Đóng'
    }

    return render_template(
        'project.html',
        projects=projects,
        stats=stats,
        active_page='project',
        filters=request.args,
        status_map=status_map
    )

# ======================
# THÊM PROJECT
# ======================
@app.route('/add_project', methods=['POST'])
def add_project():
    if 'loggedin' not in session:
        return redirect(url_for('login'))

    proj_name = request.form['proj_name']
    description = request.form['description']
    time_start = request.form['time_start']
    time_end = request.form['time_end']
    status = request.form['status']

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cursor.execute("""
        INSERT INTO projects (user_id, proj_name, description, time_start, time_end, status)
        VALUES (%s, %s, %s, %s, %s, %s)
    """, (
        session['id'],
        proj_name,
        description,
        time_start,
        time_end,
        status
    ))

    mysql.connection.commit()

    return redirect(url_for('project'))

# ======================
# SỬA PROJECT
# ======================
@app.route('/update_project', methods=['POST'])
def update_project():
    if 'loggedin' not in session:
        return redirect(url_for('login'))

    proj_id = request.form.get('proj_id')
    proj_name = request.form.get('proj_name')
    description = request.form.get('description')
    time_start = request.form.get('time_start')
    time_end = request.form.get('time_end')
    status = request.form.get('status')

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # 🔒 phân quyền
    if session['role'] == 'admin':
        cursor.execute("""
            UPDATE projects
            SET proj_name=%s, description=%s, time_start=%s, time_end=%s, status=%s
            WHERE proj_id=%s
        """, (proj_name, description, time_start, time_end, status, proj_id))
    else:
        cursor.execute("""
            UPDATE projects
            SET proj_name=%s, description=%s, time_start=%s, time_end=%s, status=%s
            WHERE proj_id=%s AND user_id=%s
        """, (proj_name, description, time_start, time_end, status, proj_id, session['id']))

    mysql.connection.commit()

    return redirect(url_for('project'))

# ======================
# XÓA PROJECT
# ======================
@app.route('/delete_project', methods=['POST'])
def delete_project():
    if 'loggedin' not in session:
        return redirect(url_for('login'))

    proj_id = request.form.get('proj_id')
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    if session['role'] == 'admin':
        cursor.execute("DELETE FROM projects WHERE proj_id = %s", (proj_id,))
    else:
        cursor.execute("""
            DELETE FROM projects 
            WHERE proj_id = %s AND user_id = %s
        """, (proj_id, session['id']))

    mysql.connection.commit()

    return redirect(url_for('project'))

# ======================
# 📥 IMPORT DATA (WORK)
# ======================

@app.route('/importdata')
def importdata():

    # ===== CHECK LOGIN =====
    if 'loggedin' not in session:
        return redirect(url_for('login'))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # ===== LẤY PARAM =====
    proj_id = request.args.get('proj_id')
    keyword = request.args.get('keyword', '').strip()

    # ===== CHECK PROJ_ID =====
    if not proj_id:
        flash("Vui lòng chọn dự án để nhập dữ liệu", "warning")
        return redirect(url_for('project'))

    # ===== LẤY PROJECT (PHÂN QUYỀN) =====
    if session['role'] == 'admin':
        cursor.execute("SELECT * FROM projects WHERE proj_id = %s", (proj_id,))
    else:
        cursor.execute("""
            SELECT * FROM projects
            WHERE proj_id = %s AND user_id = %s
        """, (proj_id, session['id']))

    project = cursor.fetchone()

    if not project:
        return "Không có quyền truy cập dự án!", 403

    # ===== LẤY WORK + SEARCH =====
    if keyword:
        cursor.execute("""
            SELECT * FROM work
            WHERE proj_id = %s AND work_name LIKE %s
        """, (proj_id, f"%{keyword}%"))
    else:
        cursor.execute("""
            SELECT * FROM work
            WHERE proj_id = %s
        """, (proj_id,))

    works = cursor.fetchall()

    # ===== TÍNH TOÁN EVM =====
    total_PV = total_EV = total_AC = 0

    for w in works:

        # ===== PV, EV, AC =====
        PV = (w.get('PV_kluong') or 0) * (w.get('PV_dongia') or 0)
        EV = (w.get('EV_kluong') or 0) * (w.get('EV_dongia') or 0)
        AC = (w.get('AC_kluong') or 0) * (w.get('AC_dongia') or 0)

        total_PV += PV
        total_EV += EV
        total_AC += AC

        # ===== GÁN LẠI =====
        w['PV'] = PV
        w['EV'] = EV
        w['AC'] = AC

        # ===== SV, CV =====
        w['SV'] = EV - PV
        w['CV'] = EV - AC

        # ===== CPI, SPI =====
        w['CPI'] = round(EV / AC, 2) if AC else 0
        w['SPI'] = round(EV / PV, 2) if PV else 0

        # ===== THỜI GIAN =====
        if w.get('time_start') and w.get('time_end'):
            w['time'] = (w['time_end'] - w['time_start']).days
        else:
            w['time'] = 0

    # ===== KPI TOÀN PROJECT =====
    BAC = project.get('BAC') or total_PV

    CV = total_EV - total_AC
    SV = total_EV - total_PV

    CPI = total_EV / total_AC if total_AC else 0
    SPI = total_EV / total_PV if total_PV else 0

    EAC = BAC / CPI if CPI else 0
    VAC = BAC - EAC

    kpi = {
        'PV': round(total_PV, 2),
        'EV': round(total_EV, 2),
        'AC': round(total_AC, 2),
        'BAC': round(BAC, 2),
        'CV': round(CV, 2),
        'SV': round(SV, 2),
        'CPI': round(CPI, 2),
        'SPI': round(SPI, 2),
        'EAC': round(EAC, 2),
        'VAC': round(VAC, 2),
    }

    # ===== STATUS MAP =====
    status_map = {
        'ke_hoach': 'Kế hoạch',
        'dang_thuc_hien': 'Đang thực hiện',
        'tam_dung': 'Tạm dừng',
        'hoan_thanh': 'Hoàn thành',
        'dong': 'Đóng'
    }

    return render_template(
        'importdata.html',
        project=project,
        works=works,
        kpi=kpi,
        status_map=status_map,
        active_page='importdata'
    )

# ======================
# THÊM CÔNG VIỆC
# ======================

@app.route('/add_work', methods=['POST'])
def add_work():

    proj_id = request.form.get('proj_id')

    PV_kluong = parse_number(request.form['PV_kluong'])
    PV_dongia = parse_number(request.form['PV_dongia'])

    EV_kluong = parse_number(request.form['EV_kluong'])
    EV_dongia = parse_number(request.form['EV_dongia'])

    AC_kluong = parse_number(request.form['AC_kluong'])
    AC_dongia = parse_number(request.form['AC_dongia'])

    start = request.form.get('time_start')
    end = request.form.get('time_end')

    # ===== LẤY THỜI GIAN DỰ ÁN =====
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cursor.execute("""
        SELECT time_start, time_end
        FROM projects
        WHERE proj_id = %s
    """, (proj_id,))

    project = cursor.fetchone()

    project_start = project['time_start']
    project_end = project['time_end']

    # ===== VALIDATE =====
    if start and end:

        start_date = datetime.strptime(start, "%Y-%m-%d").date()
        end_date = datetime.strptime(end, "%Y-%m-%d").date()

        if end_date < start_date:
            flash("Ngày kết thúc phải lớn hơn ngày bắt đầu!", "danger")
            return redirect(url_for('importdata', proj_id=proj_id))

        if start_date < project_start or end_date > project_end:
            flash(
                "Thời gian công việc phải nằm trong thời gian dự án!",
                "danger"
            )
            return redirect(url_for('importdata', proj_id=proj_id))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cursor.execute("""
        INSERT INTO work (
            proj_id, work_name,
            PV_kluong, PV_dongia,
            EV_kluong, EV_dongia,
            AC_kluong, AC_dongia,
            time_start, time_end
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        proj_id, request.form['work_name'],
        PV_kluong, PV_dongia,
        EV_kluong, EV_dongia,
        AC_kluong, AC_dongia,
        start, end
    ))

    mysql.connection.commit()

    return redirect(url_for('importdata', proj_id=proj_id))

# ======================
# UPDATE CÔNG VIỆC
# ======================
@app.route('/update_work', methods=['POST'])
def update_work():

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    work_id = request.form['work_id']
    proj_id = request.form['proj_id']
    work_name = request.form['work_name']
    PV_kluong = parse_number(request.form['PV_kluong'])
    PV_dongia = parse_number(request.form['PV_dongia'])
    EV_kluong = parse_number(request.form['EV_kluong'])
    EV_dongia = parse_number(request.form['EV_dongia'])
    AC_kluong = parse_number(request.form['AC_kluong'])
    AC_dongia = parse_number(request.form['AC_dongia'])
    time_start = request.form['time_start']
    time_end = request.form['time_end']

    # ===== LẤY THỜI GIAN DỰ ÁN =====
    cursor.execute("""
        SELECT time_start, time_end
        FROM projects
        WHERE proj_id = %s
    """, (proj_id,))

    project = cursor.fetchone()

    project_start = project['time_start']
    project_end = project['time_end']

    start_date = datetime.strptime(
        time_start,
        "%Y-%m-%d"
    ).date()

    end_date = datetime.strptime(
        time_end,
        "%Y-%m-%d"
    ).date()

    # ===== VALIDATE =====
    if end_date < start_date:
        flash(
            "Ngày kết thúc phải lớn hơn ngày bắt đầu!",
            "danger"
        )

        return redirect(url_for(
            'importdata',
            proj_id=proj_id
        ))

    if start_date < project_start or end_date > project_end:

        flash(
            "Thời gian công việc phải nằm trong thời gian dự án!",
            "danger"
        )

        return redirect(url_for(
            'importdata',
            proj_id=proj_id
        ))

    cursor.execute("""
        UPDATE work
        SET
            work_name = %s,
            PV_kluong = %s,
            PV_dongia = %s,
            EV_kluong = %s,
            EV_dongia = %s,
            AC_kluong = %s,
            AC_dongia = %s,
            time_start = %s,
            time_end = %s

        WHERE work_id = %s
    """, (
        work_name,
        PV_kluong,
        PV_dongia,
        EV_kluong,
        EV_dongia,
        AC_kluong,
        AC_dongia,
        time_start,
        time_end,
        work_id
    ))

    mysql.connection.commit()

    return redirect(url_for(
        'importdata',
        proj_id=proj_id
    ))

# ======================
# XÓA CÔNG VIỆC
# ======================
@app.route('/delete_work/<int:work_id>')
def delete_work(work_id):

    proj_id = request.args.get('proj_id')

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cursor.execute("""
        DELETE FROM work WHERE work_id = %s
    """, (work_id,))

    mysql.connection.commit()

    return redirect(url_for(
        'importdata',
        proj_id=proj_id
    ))

# ======================
# TẢI FILE EXCEL MẪU
# ======================
@app.route('/download_excel_template')
def download_excel_template():

    if 'loggedin' not in session:
        return redirect(url_for('login'))

    return send_file(
        'static/sample/mau_danh_sach_cong_viec.xlsx',
        as_attachment=True,
        download_name='mau_danh_sach_cong_viec.xlsx'
    )

# ======================
# IMPORT FILE EXCEL
# ======================
@app.route('/import_excel', methods=['POST'])
def import_excel():

    if 'loggedin' not in session:
        return redirect(url_for('login'))

    file = request.files.get('excel_file')
    proj_id = request.form.get('proj_id')

    if not file or file.filename == '':
        flash("Vui lòng chọn file excel!", "danger")
        return redirect(url_for('importdata', proj_id=proj_id))

    # ===== CHECK ĐUÔI FILE =====
    if not file.filename.endswith(('.xlsx', '.xls')):
        flash("File không đúng định dạng excel!", "danger")
        return redirect(url_for('importdata', proj_id=proj_id))

    try:

        # ===== ĐỌC FILE =====
        df = pd.read_excel(file, engine='openpyxl')

        # ===== CHECK CỘT =====
        required_columns = [
            'Tên công việc',
            'Ngày bắt đầu',
            'Ngày hoàn thành',
            'Khối lượng PV',
            'Đơn giá PV (VNĐ)',
            'Khối lượng EV',
            'Đơn giá EV (VNĐ)',
            'Khối lượng AC',
            'Đơn giá AC (VNĐ)'
        ]

        missing_cols = []

        for col in required_columns:
            if col not in df.columns:
                missing_cols.append(col)

        if missing_cols:
            flash(
                "Thiếu cột: " + ", ".join(missing_cols),
                "danger"
            )
            return redirect(url_for('importdata', proj_id=proj_id))

        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

        # ===== LẤY THỜI GIAN DỰ ÁN =====
        cursor.execute("""
            SELECT time_start, time_end
            FROM projects
            WHERE proj_id = %s
        """, (proj_id,))

        project = cursor.fetchone()

        project_start = project['time_start']
        project_end = project['time_end']

        total = len(df)
        success = 0
        error = 0

        for index, row in df.iterrows():

            try:

                # ===== TÊN CÔNG VIỆC =====
                work_name = str(row['Tên công việc']).strip()

                if work_name == "" or work_name.lower() == "nan":
                    raise Exception("Tên công việc bị trống")

                # ===== DATE =====
                time_start = pd.to_datetime(
                    row['Ngày bắt đầu'],
                    errors='raise'
                ).date()

                time_end = pd.to_datetime(
                    row['Ngày hoàn thành'],
                    errors='raise'
                ).date()

                # ===== VALIDATE DATE =====
                if time_end < time_start:
                    raise Exception(
                        "Ngày hoàn thành nhỏ hơn ngày bắt đầu"
                    )

                # ===== CHECK TRONG DỰ ÁN =====
                if time_start < project_start or time_end > project_end:
                    raise Exception(
                        "Thời gian công việc nằm ngoài thời gian dự án"
                    )

                # ===== NUMBER FUNCTION =====
                def to_number(value):

                    if pd.isna(value):
                        return 0

                    value = str(value).replace(',', '').strip()

                    return float(value)

                # ===== CONVERT NUMBER =====
                PV_kluong = to_number(row['Khối lượng PV'])
                PV_dongia = to_number(row['Đơn giá PV (VNĐ)'])

                EV_kluong = to_number(row['Khối lượng EV'])
                EV_dongia = to_number(row['Đơn giá EV (VNĐ)'])

                AC_kluong = to_number(row['Khối lượng AC'])
                AC_dongia = to_number(row['Đơn giá AC (VNĐ)'])

                # ===== VALIDATE ÂM =====
                numbers = [
                    PV_kluong,
                    PV_dongia,
                    EV_kluong,
                    EV_dongia,
                    AC_kluong,
                    AC_dongia
                ]

                for num in numbers:
                    if num < 0:
                        raise Exception("Không được nhập số âm")

                # ===== INSERT =====
                cursor.execute("""
                    INSERT INTO work (

                        proj_id,
                        work_name,

                        time_start,
                        time_end,

                        PV_kluong,
                        PV_dongia,

                        EV_kluong,
                        EV_dongia,

                        AC_kluong,
                        AC_dongia

                    )

                    VALUES (

                        %s, %s,
                        %s, %s,

                        %s, %s,
                        %s, %s,
                        %s, %s
                    )
                """, (

                    proj_id,
                    work_name,

                    time_start,
                    time_end,

                    PV_kluong,
                    PV_dongia,

                    EV_kluong,
                    EV_dongia,

                    AC_kluong,
                    AC_dongia
                ))

                success += 1

            except Exception as e:

                print(f"Lỗi dòng {index + 2}: {e}")

                error += 1

        mysql.connection.commit()

        flash(
            f"""
            Import hoàn tất!
            Tổng dòng: {total}
            | Thành công: {success}
            | Lỗi: {error}
            """,
            "success"
        )

    except Exception as e:

        print(e)

        flash(
            "File excel không hợp lệ hoặc sai cấu trúc!",
            "danger"
        )

    return redirect(url_for('importdata', proj_id=proj_id))

# ======================
# SCURVE
# ======================
@app.route('/scurve')
def scurve():

    if 'loggedin' not in session:
        return redirect(url_for('login'))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # ===== LẤY PROJECT =====

    if session['role'] == 'admin':
        cursor.execute("SELECT * FROM projects")
    else:
        cursor.execute("""
            SELECT * FROM projects
            WHERE user_id = %s
        """, (session['id'],))

    projects = cursor.fetchall()

    # ===== PROJECT ĐƯỢC CHỌN =====

    selected_proj_id = request.args.get('proj_id')

    if not selected_proj_id:

        return render_template(
            'scurve.html',
            projects=projects,
            selected_proj_id=None,
            active_page='scurve'
        )

    # ===== LẤY WORK =====

    cursor.execute("""
        SELECT *
        FROM work
        WHERE proj_id = %s
        ORDER BY time_start
    """, (selected_proj_id,))

    works = cursor.fetchall()

    # ===== DATA CHART =====

    labels = []
    pv_data = []
    ev_data = []
    ac_data = []

    total_PV = 0
    total_EV = 0
    total_AC = 0

    for w in works:

        labels.append(w['work_name'])

        PV = (w['PV_kluong'] or 0) * (w['PV_dongia'] or 0)
        EV = (w['EV_kluong'] or 0) * (w['EV_dongia'] or 0)
        AC = (w['AC_kluong'] or 0) * (w['AC_dongia'] or 0)

        total_PV += PV
        total_EV += EV
        total_AC += AC

        pv_data.append(total_PV)
        ev_data.append(total_EV)
        ac_data.append(total_AC)

    SPI = round(total_EV / total_PV, 2) if total_PV else 0
    CPI = round(total_EV / total_AC, 2) if total_AC else 0

    # ===== PROJECT =====

    cursor.execute("""
        SELECT *
        FROM projects
        WHERE proj_id = %s
    """, (selected_proj_id,))

    project = cursor.fetchone()

    return render_template(
        'scurve.html',

        active_page='scurve',

        projects=projects,
        project=project,

        selected_proj_id=int(selected_proj_id),

        labels=labels,
        pv_data=pv_data,
        ev_data=ev_data,
        ac_data=ac_data,

        total_PV=round(total_PV, 2),
        total_EV=round(total_EV, 2),
        total_AC=round(total_AC, 2),

        SPI=SPI,
        CPI=CPI
    )

# ======================
# REPORT
# ======================
@app.route('/report')
def report():

    if 'loggedin' not in session:
        return redirect(url_for('login'))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # ===== DANH SÁCH PROJECT =====
    if session['role'] == 'admin':
        cursor.execute("SELECT * FROM projects")
    else:
        cursor.execute("""
            SELECT * FROM projects
            WHERE user_id = %s
        """, (session['id'],))

    projects = cursor.fetchall()

    proj_id = request.args.get('proj_id')

    project = None
    works = []
    kpi = {}

    status_map = {
        'ke_hoach': 'Kế hoạch',
        'dang_thuc_hien': 'Đang thực hiện',
        'tam_dung': 'Tạm dừng',
        'hoan_thanh': 'Hoàn thành',
        'dong': 'Đóng'
    }

    if proj_id:

        cursor.execute("""
            SELECT * FROM projects
            WHERE proj_id = %s
        """, (proj_id,))

        project = cursor.fetchone()

        cursor.execute("""
            SELECT * FROM work
            WHERE proj_id = %s
        """, (proj_id,))

        works = cursor.fetchall()

        total_PV = total_EV = total_AC = 0

        for w in works:

            w['PV'] = w['PV_kluong'] * w['PV_dongia']
            w['EV'] = w['EV_kluong'] * w['EV_dongia']
            w['AC'] = w['AC_kluong'] * w['AC_dongia']

            w['SV'] = w['EV'] - w['PV']
            w['CV'] = w['EV'] - w['AC']

            w['SPI'] = round(
                w['EV'] / w['PV'], 2
            ) if w['PV'] else 0

            w['CPI'] = round(
                w['EV'] / w['AC'], 2
            ) if w['AC'] else 0

            total_PV += w['PV']
            total_EV += w['EV']
            total_AC += w['AC']

        CPI = total_EV / total_AC if total_AC else 0
        SPI = total_EV / total_PV if total_PV else 0

        BAC = total_PV
        EAC = BAC / CPI if CPI else 0

        kpi = {
            'PV': total_PV,
            'EV': total_EV,
            'AC': total_AC,
            'CPI': round(CPI, 2),
            'SPI': round(SPI, 2),
            'EAC': round(EAC, 2)
        }

    return render_template(
        'report.html',

        projects=projects,
        project=project,
        works=works,
        kpi=kpi,

        status_map=status_map,

        selected_proj_id=proj_id,

        active_page='report'
    )

# =====================================================
# TẠO FILE PDF REPORT
# =====================================================
def build_pdf_report(proj_id):

    # ===== FONT =====
    pdfmetrics.registerFont(
        TTFont(
            'Arial',
            'static/fonts/arial.ttf'
        )
    )

    # ===== BUFFER PDF =====
    pdf_buffer = BytesIO()

    # ===== GET DATA =====
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cursor.execute("""
        SELECT *
        FROM projects
        WHERE proj_id = %s
    """, (proj_id,))

    project = cursor.fetchone()

    cursor.execute("""
        SELECT *
        FROM work
        WHERE proj_id = %s
    """, (proj_id,))

    works = cursor.fetchall()

    # ===== KPI =====
    total_PV = 0
    total_EV = 0
    total_AC = 0

    for w in works:

        PV = (w['PV_kluong'] or 0) * (w['PV_dongia'] or 0)
        EV = (w['EV_kluong'] or 0) * (w['EV_dongia'] or 0)
        AC = (w['AC_kluong'] or 0) * (w['AC_dongia'] or 0)

        w['PV'] = PV
        w['EV'] = EV
        w['AC'] = AC

        w['SV'] = EV - PV
        w['CV'] = EV - AC

        w['SPI'] = round(EV / PV, 2) if PV else 0
        w['CPI'] = round(EV / AC, 2) if AC else 0

        total_PV += PV
        total_EV += EV
        total_AC += AC

    BAC = total_PV

    SV = total_EV - total_PV
    CV = total_EV - total_AC

    SPI = round(total_EV / total_PV, 2) if total_PV else 0
    CPI = round(total_EV / total_AC, 2) if total_AC else 0

    CSI = round(SPI * CPI, 2)

    EAC = round(BAC / CPI, 2) if CPI else 0
    ETC = round(EAC - total_AC, 2)
    VAC = round(BAC - EAC, 2)

    TCPI = round(
        (BAC - total_EV) / (BAC - total_AC),
        2
    ) if (BAC - total_AC) != 0 else 0

    # ===== PDF =====
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=18
    )

    styles = getSampleStyleSheet()

    style = styles['Normal']
    style.fontName = 'Arial'
    style.fontSize = 11
    style.leading = 20

    title_h1_style = styles['Heading1']
    title_h1_style.fontName = 'Arial'
    title_h1_style.alignment = 1

    title_h2_style = styles['Heading2']
    title_h2_style.fontName = 'Arial'

    elements = []

    # ======================
    # TITLE
    # ======================

    elements.append(
        Paragraph(
            "<b>BÁO CÁO THEO DÕI CHI PHÍ VÀ TIẾN ĐỘ DỰ ÁN</b>",
            title_h1_style
        )
    )

    elements.append(Spacer(1, 20))

    # ======================
    # THÔNG TIN DỰ ÁN
    # ======================

    elements.append(
        Paragraph(
            "<b>1. THÔNG TIN DỰ ÁN</b><br/>",
            title_h2_style
        )
    )

    status_map = {
        'ke_hoach': 'Kế hoạch',
        'dang_thuc_hien': 'Đang thực hiện',
        'tam_dung': 'Tạm dừng',
        'hoan_thanh': 'Hoàn thành',
        'dong': 'Đóng'
    }

    project_info = f"""
    <b>Tên dự án:</b> {project['proj_name']}<br/>
    <b>Ngày bắt đầu:</b> {project['time_start']}<br/>
    <b>Ngày kết thúc:</b> {project['time_end']}<br/>
    <b>Trạng thái:</b> {status_map.get(project['status'], '')}
    """

    elements.append(
        Paragraph(project_info, style)
    )

    elements.append(Spacer(1, 20))

    # ======================
    # KPI
    # ======================

    elements.append(
        Paragraph(
            "<b>2. TỔNG QUAN KPI</b><br/>",
            title_h2_style
        )
    )

    kpi_text = f"""
    <b>PV:</b> {total_PV:,.0f} VNĐ<br/>
    <b>EV:</b> {total_EV:,.0f} VNĐ<br/>
    <b>AC:</b> {total_AC:,.0f} VNĐ<br/>
    <b>BAC:</b> {BAC:,.0f} VNĐ<br/><br/>

    <b>SV:</b> {SV:,.0f} VNĐ<br/>
    <b>CV:</b> {CV:,.0f} VNĐ<br/><br/>

    <b>SPI:</b> {SPI}<br/>
    <b>CPI:</b> {CPI}<br/>
    <b>CSI:</b> {CSI}<br/><br/>

    <b>EAC:</b> {EAC:,.0f} VNĐ<br/>
    <b>ETC:</b> {ETC:,.0f} VNĐ<br/>
    <b>VAC:</b> {VAC:,.0f} VNĐ<br/>
    <b>TCPI:</b> {TCPI}
    """

    elements.append(
        Paragraph(kpi_text, style)
    )

    elements.append(Spacer(1, 20))

    # ======================
    # ĐÁNH GIÁ
    # ======================

    elements.append(
        Paragraph(
            "<b>3. ĐÁNH GIÁ</b><br/>",
            title_h2_style
        )
    )

    evaluate = ""

    if SPI < 1:
        evaluate += "• Dự án đang chậm tiến độ.<br/>"
    else:
        evaluate += "• Dự án đang đúng tiến độ.<br/>"

    if CPI < 1:
        evaluate += "• Dự án đang vượt ngân sách.<br/>"
    else:
        evaluate += "• Dự án đang tiết kiệm chi phí.<br/>"

    elements.append(
        Paragraph(evaluate, style)
    )

    elements.append(Spacer(1, 20))

    # ======================
    # BẢNG CÔNG VIỆC
    # ======================

    elements.append(
        Paragraph(
            "<b>4. CHI TIẾT CÔNG VIỆC</b><br/>",
            title_h2_style
        )
    )

    data = [[
        'STT',
        'Công việc',
        'Ngày bắt đầu',
        'Ngày kết thúc',
        'PV',
        'EV',
        'AC',
        'SPI',
        'CPI'
    ]]

    for i, w in enumerate(works, start=1):

        data.append([
            i,
            w['work_name'],
            str(w['time_start']),
            str(w['time_end']),
            f"{w['PV']:,.0f}",
            f"{w['EV']:,.0f}",
            f"{w['AC']:,.0f}",
            str(w['SPI']),
            str(w['CPI'])
        ])

    table = Table(data, repeatRows=1)

    table.setStyle(TableStyle([

        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),

        ('FONTNAME', (0,0), (-1,-1), 'Arial'),

        ('GRID', (0,0), (-1,-1), 1, colors.black),

        ('ALIGN', (0,0), (-1,-1), 'CENTER'),

        ('BOTTOMPADDING', (0,0), (-1,0), 10),

    ]))

    elements.append(table)

    elements.append(Spacer(1, 30))

    # ======================
    # KẾT LUẬN
    # ======================

    elements.append(
        Paragraph(
            "<b>5. KẾT LUẬN</b><br/>",
            title_h2_style
        )
    )

    conclusion = """
    • Cần tiếp tục theo dõi tiến độ và chi phí.<br/>
    • Tăng cường kiểm soát nguồn lực thi công.<br/>
    • Theo dõi CPI và SPI định kỳ.
    """

    elements.append(
        Paragraph(conclusion, style)
    )

    # ===== BUILD PDF =====
    doc.build(elements)

    pdf_buffer.seek(0)

    return pdf_buffer, project

# ======================
# EXPORT PDF
# ======================
@app.route('/export_pdf/<int:proj_id>')
def export_pdf(proj_id):

    pdf_buffer, project = build_pdf_report(proj_id)

    project_name = project['proj_name'].replace(" ", "_")

    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name=f"Bao_cao_{project_name}.pdf",
        mimetype='application/pdf'
    )
# ======================
# EXPORT EXCEL
# ======================
@app.route('/export_excel/<int:proj_id>')
def export_excel(proj_id):

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cursor.execute("""
        SELECT *
        FROM projects
        WHERE proj_id = %s
    """, (proj_id,))

    project = cursor.fetchone()

    cursor.execute("""
        SELECT *
        FROM work
        WHERE proj_id = %s
    """, (proj_id,))

    works = cursor.fetchall()

    project_name = project['proj_name'].replace(" ", "_")

    data = []

    total_PV = 0
    total_EV = 0
    total_AC = 0

    for i, w in enumerate(works, start=1):

        PV = (w['PV_kluong'] or 0) * (w['PV_dongia'] or 0)
        EV = (w['EV_kluong'] or 0) * (w['EV_dongia'] or 0)
        AC = (w['AC_kluong'] or 0) * (w['AC_dongia'] or 0)

        SV = EV - PV
        CV = EV - AC

        SPI = round(EV / PV, 2) if PV else 0
        CPI = round(EV / AC, 2) if AC else 0

        total_PV += PV
        total_EV += EV
        total_AC += AC

        data.append({

            'STT': i,
            'Công việc': w['work_name'],
            'Ngày bắt đầu': w['time_start'],
            'Ngày kết thúc': w['time_end'],

            'PV': PV,
            'EV': EV,
            'AC': AC,

            'SV': SV,
            'CV': CV,

            'SPI': SPI,
            'CPI': CPI
        })

    # =================================================
    # TỔNG
    # =================================================

    data.append({

        'STT': 'TỔNG',
        'Công việc': '',

        'Ngày bắt đầu': project['time_start'],
        'Ngày kết thúc': project['time_end'],

        'PV': total_PV,
        'EV': total_EV,
        'AC': total_AC,

        'SV': total_EV - total_PV,
        'CV': total_EV - total_AC,

        'SPI': round(total_EV / total_PV, 2)
        if total_PV else 0,

        'CPI': round(total_EV / total_AC, 2)
        if total_AC else 0
    })

    df = pd.DataFrame(data)

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine='openpyxl'
    ) as writer:

        df.to_excel(
            writer,
            sheet_name='EVM Report',
            startrow=3,
            index=False
        )

        workbook = writer.book

        worksheet = writer.sheets['EVM Report']

        # =================================================
        # TITLE
        # =================================================

        worksheet.merge_cells('A1:K1')

        title_cell = worksheet['A1']

        title_cell.value = (
            "BÁO CÁO THEO DÕI CHI PHÍ "
            "VÀ TIẾN ĐỘ DỰ ÁN"
        )

        title_cell.font = Font(
            bold=True,
            size=16,
            color='000000'
        )

        title_cell.alignment = Alignment(
            horizontal='center',
            vertical='center'
        )

        # =================================================
        # HEADER
        # =================================================

        header_fill = PatternFill(
            start_color='1F4E78',
            end_color='1F4E78',
            fill_type='solid'
        )

        thin = Side(style='thin')

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        header_row = 4

        for cell in worksheet[header_row]:

            cell.font = Font(
                bold=True,
                color='FFFFFF'
            )

            cell.fill = header_fill

            cell.alignment = Alignment(
                horizontal='center',
                vertical='center'
            )

            cell.border = border

        # =================================================
        # STT STYLE
        # =================================================

        for row in worksheet.iter_rows(
            min_row=5,
            max_row=worksheet.max_row
        ):

            for cell in row:

                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center'
                )

                cell.border = border

        # =================================================
        # FORMAT NUMBER
        # =================================================

        number_columns = ['E', 'F', 'G', 'H', 'I']

        for col in number_columns:

            for row in range(
                5,
                worksheet.max_row + 1
            ):

                worksheet[
                    f'{col}{row}'
                ].number_format = '#,##0'

        # =================================================
        # FORMAT DECIMAL
        # =================================================

        decimal_columns = ['J', 'K']

        for col in decimal_columns:

            for row in range(
                5,
                worksheet.max_row + 1
            ):

                worksheet[
                    f'{col}{row}'
                ].number_format = '0.00'

        # =================================================
        # TỰ ĐỘNG GIÃN CỘT
        # =================================================

        for column_cells in worksheet.columns:

            max_length = 0

            column = column_cells[0].column

            for cell in column_cells:

                try:

                    if len(str(cell.value)) > max_length:

                        max_length = len(
                            str(cell.value)
                        )

                except:
                    pass

            adjusted_width = max_length + 5

            worksheet.column_dimensions[
                get_column_letter(column)
            ].width = adjusted_width


        total_row = worksheet.max_row

        worksheet.merge_cells(

            start_row=total_row,
            start_column=1,

            end_row=total_row,
            end_column=2
        )

        total_cell = worksheet.cell(
            row=total_row,
            column=1
        )

        total_cell.font = Font(
            bold=True,
            color='FFFFFF'
        )

        total_cell.fill = PatternFill(
            start_color='C00000',
            end_color='C00000',
            fill_type='solid'
        )

        total_cell.alignment = Alignment(
            horizontal='center',
            vertical='center'
        )

        worksheet.freeze_panes = 'A5'

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )


    output.seek(0)

    return send_file(

        output,

        as_attachment=True,

        download_name=(
            f'Bao_cao_{project_name}.xlsx'
        ),

        mimetype=(
            'application/'
            'vnd.openxmlformats-officedocument.'
            'spreadsheetml.sheet'
        )
    )
# ======================
# GỬI REPORT EMAIL
# ======================
@app.route('/send_report_email/<int:proj_id>')
def send_report_email(proj_id):

    if 'loggedin' not in session:
        return redirect(url_for('login'))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # ===== USER =====
    cursor.execute("""
        SELECT email, fullname
        FROM users
        WHERE user_id = %s
    """, (session['id'],))

    user = cursor.fetchone()

    # ===== BUILD PDF =====
    pdf_buffer, project = build_pdf_report(proj_id)

    try:

        msg = Message(
            subject=f"Báo cáo EVM - {project['proj_name']}",
            recipients=[user['email']]
        )

        msg.body = f"""
Xin chào {user['fullname']},

Hệ thống gửi file đính kèm là báo cáo EVM của dự án:
{project['proj_name']}

Trân trọng,
Hệ thống EVM
"""

        # ===== ATTACH PDF =====
        msg.attach(
            filename=f"Bao_cao_{project['proj_name']}.pdf",
            content_type='application/pdf',
            data=pdf_buffer.read()
        )

        mail.send(msg)

        flash(
            "Gửi email thành công!",
            "success"
        )

    except Exception as e:

        print(e)

        flash(
            "Gửi email thất bại!",
            "danger"
        )

    return redirect(url_for(
        'report',
        proj_id=proj_id
    ))
# ======================
# ACCOUNT
# ======================
@app.route('/account')
def account():

    if 'loggedin' not in session:
        return redirect(url_for('login'))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # USER
    if session['role'] == 'user':

        cursor.execute("""
            SELECT * FROM users
            WHERE user_id = %s
        """, (session['id'],))

        account = cursor.fetchone()

        return render_template(
            'account.html',
            account=account,
            active_page='account'
        )

    # ADMIN
    else:

        keyword = request.args.get('keyword', '').strip()

        if keyword:
            cursor.execute("""
                SELECT * FROM users
                WHERE fullname LIKE %s
                OR username LIKE %s
                OR email LIKE %s
            """, (
                f"%{keyword}%",
                f"%{keyword}%",
                f"%{keyword}%"
            ))
        else:
            cursor.execute("""
                SELECT * FROM users
            """)

        accounts = cursor.fetchall()

        return render_template(
            'account.html',
            accounts=accounts,
            active_page='account'
        )
    
# ======================
# ADD ACCOUNT
# ======================
@app.route('/add_account', methods=['POST'])
def add_account():

    if 'loggedin' not in session:
        return redirect(url_for('login'))

    if session['role'] != 'admin':
        return "Không có quyền!", 403

    form = request.form

    fullname = form.get('fullname', '').strip()
    email = form.get('email', '').strip()
    phone = form.get('phone', '').strip()
    username = form.get('username', '').strip()
    password = form.get('password', '').strip()
    role = form.get('role')

    form_data = {
        "fullname": fullname,
        "email": email,
        "phone": phone,
        "username": username,
        "role": role
    }

    def fail(msg):
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute("SELECT * FROM users")
        accounts = cursor.fetchall()

        flash(msg, "warning")

        return render_template(
            "account.html",
            accounts=accounts,
            form_data=form_data,
            open_modal=True,
            active_page='account'
        )

    # ===== VALIDATE =====
    if not all([fullname, email, phone, username, password]):
        return fail("Vui lòng nhập đầy đủ thông tin!")

    if not phone.isdigit():
        return fail("Số điện thoại không hợp lệ!")

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cursor.execute("SELECT 1 FROM users WHERE email=%s", (email,))
    if cursor.fetchone():
        return fail("Email đã tồn tại!")

    cursor.execute("SELECT 1 FROM users WHERE username=%s", (username,))
    if cursor.fetchone():
        return fail("Username đã tồn tại!")

    # ===== INSERT =====
    cursor.execute("""
        INSERT INTO users
        (fullname, email, phone, username, password, role)
        VALUES (%s, %s, %s, %s, %s, %s)
    """, (fullname, email, phone, username, password, role))

    mysql.connection.commit()

    flash("Thêm tài khoản thành công!", "success")

    return redirect(url_for('account'))

# ======================
# UPDATE ACCOUNT ADMIN
# ======================
@app.route('/update_account_admin', methods=['POST'])
def update_account_admin():

    if 'loggedin' not in session:
        return redirect(url_for('login'))

    if session['role'] != 'admin':
        return "Không có quyền!", 403

    user_id = request.form.get('user_id')

    fullname = request.form.get('fullname', '').strip()
    email = request.form.get('email', '').strip()
    phone = request.form.get('phone', '').strip()
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    role = request.form.get('role')

    # ===== CHECK RỖNG =====
    if not all([fullname, email, phone, username, password, role]):
        flash("Vui lòng nhập đầy đủ thông tin!", "warning")
        return redirect(url_for('account'))

    # ===== CHECK PHONE =====
    if not phone.isdigit():
        flash("Số điện thoại không hợp lệ!", "warning")
        return redirect(url_for('account'))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # ===== CHECK EMAIL TRÙNG =====
    cursor.execute("""
        SELECT * FROM users
        WHERE email = %s
        AND user_id != %s
    """, (email, user_id))

    if cursor.fetchone():
        flash("Email đã tồn tại!", "warning")
        return redirect(url_for('account'))

    # ===== CHECK USERNAME TRÙNG =====
    cursor.execute("""
        SELECT * FROM users
        WHERE username = %s
        AND user_id != %s
    """, (username, user_id))

    if cursor.fetchone():
        flash("Username đã tồn tại!", "warning")
        return redirect(url_for('account'))

    # ===== UPDATE =====
    cursor.execute("""
        UPDATE users
        SET
            fullname = %s,
            email = %s,
            phone = %s,
            username = %s,
            password = %s,
            role = %s
        WHERE user_id = %s
    """, (
        fullname,
        email,
        phone,
        username,
        password,
        role,
        user_id
    ))

    mysql.connection.commit()

    flash("Cập nhật tài khoản thành công!", "success")

    return redirect(url_for('account'))

# ======================
# UPDATE ACCOUNT USER
# ======================
@app.route('/update_account', methods=['POST'])
def update_account():

    if 'loggedin' not in session:
        return redirect(url_for('login'))

    fullname = request.form.get('fullname').strip()
    email = request.form.get('email').strip()
    phone = request.form.get('phone').strip()
    username = request.form.get('username').strip()
    password = request.form.get('password').strip()

    # ===== VALIDATE RỖNG =====
    if not fullname or not email or not phone or not username or not password:
        flash("Vui lòng nhập đầy đủ thông tin!", "warning")
        return redirect(url_for('account'))

    # ===== VALIDATE PHONE =====
    if not phone.isdigit():
        flash("Số điện thoại chỉ được chứa số!", "warning")
        return redirect(url_for('account'))

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # ===== CHECK EMAIL TRÙNG =====
    cursor.execute("""
        SELECT * FROM users
        WHERE email = %s AND user_id != %s
    """, (email, session['id']))

    email_exist = cursor.fetchone()

    if email_exist:
        flash("Email đã tồn tại!", "warning")
        return redirect(url_for('account'))

# ===== CHECK USERNAME TRÙNG =====
    cursor.execute("""
        SELECT * FROM users
        WHERE username = %s AND user_id != %s
    """, (username, session['id']))

    username_exist = cursor.fetchone()

    if username_exist:
        flash("Username đã tồn tại!", "warning")
        return redirect(url_for('account'))

    # ===== UPDATE =====
    cursor.execute("""
        UPDATE users
        SET fullname = %s,
            email = %s,
            phone = %s,
            username = %s,
            password = %s
        WHERE user_id = %s
    """, (
        fullname,
        email,
        phone,
        username,
        password,
        session['id']
    ))

    mysql.connection.commit()

    flash("Cập nhật thông tin thành công!", "success")

    return redirect(url_for('account'))

# ======================
# DELETE ACCOUNT
# ======================
@app.route('/delete_account/<int:user_id>')
def delete_account(user_id):

    if 'loggedin' not in session:
        return redirect(url_for('login'))

    if session['role'] != 'admin':
        return "Không có quyền!", 403

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cursor.execute("""
        DELETE FROM users
        WHERE user_id = %s
    """, (user_id,))

    mysql.connection.commit()

    flash("Xóa tài khoản thành công!", "success")

    return redirect(url_for('account'))

# =========================
# 🚪 Logout
# =========================
@app.route('/logout')
def logout():
    session.pop('loggedin', None)
    session.pop('id', None)
    session.pop('username', None)
    session.pop('role', None)

    return redirect(url_for('login'))

# =========================
# 🚀 Run app
# =========================
if __name__ == '__main__':
    app.run(debug=True)